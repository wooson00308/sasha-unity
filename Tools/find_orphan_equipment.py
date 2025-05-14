print("[DEBUG] Script execution started...") # Added for very basic check
import openpyxl
from pathlib import Path

# --- Configuration ---
# Path to the Excel file, relative to the project root
EXCEL_FILE_NAME = "AF_Data.xlsx"
ASSETS_DIR_NAME = "Assets"
AF_DATA_SUBDIR = "AF/Data"

# Sheet names
ASSEMBLIES_SHEET = "AF_Assemblies"
PARTS_SHEET = "Parts"
WEAPONS_SHEET = "Weapons"

# Column letters in AF_Assemblies for equipment IDs
# Ensure these are correct if the Excel structure changes!
# Part IDs in AF_Assemblies: Head(F), Body(G), ArmL(H), ArmR(I), Legs(J), Backpack(K)
# Weapon IDs in AF_Assemblies: Weapon1(L), Weapon2(M)
# (Indices: F=5, G=6, H=7, I=8, J=9, K=10, L=11, M=12)
ASSEMBLY_PART_ID_COL_INDICES = [5, 6, 7, 8, 9, 10]  # F to K
ASSEMBLY_WEAPON_ID_COL_INDICES = [11, 12]           # L to M

# Column letter for ID in Parts and Weapons sheets (usually 'A')
# (Index: A=0)
MASTER_ID_COL_INDEX = 0
# --- End Configuration ---

def get_script_directory():
    """Gets the directory where the script is located."""
    return Path(__file__).resolve().parent

def get_project_root():
    """Assumes the script is in ProjectRoot/Tools."""
    return get_script_directory().parent

def get_excel_file_path():
    """Constructs the full path to the Excel file."""
    project_root = get_project_root()
    return project_root / ASSETS_DIR_NAME / AF_DATA_SUBDIR / EXCEL_FILE_NAME

def get_used_equipment_ids(workbook):
    """Reads AF_Assemblies sheet and returns a set of all used PartIDs and WeaponIDs."""
    try:
        sheet = workbook[ASSEMBLIES_SHEET]
    except KeyError:
        print(f"Error: Sheet named '{ASSEMBLIES_SHEET}' not found in the workbook.")
        return None
        
    used_ids = set()
    
    for row_cells in sheet.iter_rows(min_row=2): # min_row=2 to skip header
        for col_idx in ASSEMBLY_PART_ID_COL_INDICES + ASSEMBLY_WEAPON_ID_COL_INDICES:
            if col_idx < len(row_cells): # Ensure column index is within row bounds
                cell_value = row_cells[col_idx].value
                if cell_value and str(cell_value).strip(): # Check for non-empty, non-None values
                    used_ids.add(str(cell_value).strip())
            else:
                # This case should ideally not happen if Excel structure is consistent
                # print(f"Warning: Row {row_cells[0].row} is shorter than expected, missing column index {col_idx}.")
                pass
    return used_ids

def get_defined_ids_from_sheet(workbook, sheet_name, id_column_index):
    """Helper function to read all unique IDs from a specific column in a sheet."""
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        print(f"Error: Sheet named '{sheet_name}' not found in the workbook.")
        return None

    defined_ids = set()
    for row_cells in sheet.iter_rows(min_row=2): # min_row=2 to skip header
        if id_column_index < len(row_cells):
            cell_value = row_cells[id_column_index].value
            if cell_value and str(cell_value).strip():
                defined_ids.add(str(cell_value).strip())
        else:
            # print(f"Warning: Row {row_cells[0].row} in sheet '{sheet_name}' is shorter than expected.")
            pass
    return defined_ids

def main():
    """Main function to find and print orphan equipment."""
    print("[DEBUG] main() called.")
    excel_file_path = get_excel_file_path()
    print(f"[DEBUG] Excel file path determined as: {excel_file_path}")

    if not excel_file_path.exists():
        print(f"[DEBUG] Excel file does NOT exist at the path.")
        print(f"Error: Excel file not found at '{excel_file_path}'")
        print("Please ensure the Excel file path configuration at the top of the script is correct.")
        return
    print(f"[DEBUG] Excel file confirmed to exist at the path.")

    try:
        print("[DEBUG] Attempting to load workbook...")
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
        print("[DEBUG] Workbook loaded successfully.")
    except Exception as e:
        print(f"[DEBUG] Exception caught while loading workbook: {e}")
        print(f"Error loading Excel workbook: {e}")
        print("Ensure 'openpyxl' is installed (pip install openpyxl) and the file is a valid Excel file.")
        return

    print(f"Analyzing Excel file: {excel_file_path.name}\n")

    # Get all used part and weapon IDs from AF_Assemblies
    print("[DEBUG] Calling get_used_equipment_ids...")
    used_equipment_ids = get_used_equipment_ids(workbook)
    if used_equipment_ids is None:
        print("[DEBUG] get_used_equipment_ids returned None. Exiting.")
        return # Error already printed
    print(f"[DEBUG] used_equipment_ids count: {len(used_equipment_ids)}")

    # Get all defined PartIDs
    print("[DEBUG] Calling get_defined_ids_from_sheet for Parts...")
    defined_part_ids = get_defined_ids_from_sheet(workbook, PARTS_SHEET, MASTER_ID_COL_INDEX)
    if defined_part_ids is None:
        print("[DEBUG] get_defined_ids_from_sheet for Parts returned None. Exiting.")
        return # Error already printed
    print(f"[DEBUG] defined_part_ids count: {len(defined_part_ids)}")
        
    # Get all defined WeaponIDs
    print("[DEBUG] Calling get_defined_ids_from_sheet for Weapons...")
    defined_weapon_ids = get_defined_ids_from_sheet(workbook, WEAPONS_SHEET, MASTER_ID_COL_INDEX)
    if defined_weapon_ids is None:
        print("[DEBUG] get_defined_ids_from_sheet for Weapons returned None. Exiting.")
        return # Error already printed
    print(f"[DEBUG] defined_weapon_ids count: {len(defined_weapon_ids)}")

    # Find orphan parts
    print("[DEBUG] Calculating orphan parts...")
    orphan_parts = defined_part_ids - used_equipment_ids
    # Find orphan weapons
    print("[DEBUG] Calculating orphan weapons...")
    orphan_weapons = defined_weapon_ids - used_equipment_ids

    print("--- Orphan Parts (Defined in 'Parts' but not used in 'AF_Assemblies') ---")
    if orphan_parts:
        for part_id in sorted(list(orphan_parts)):
            print(part_id)
    else:
        print("No orphan parts found.")

    print("\n--- Orphan Weapons (Defined in 'Weapons' but not used in 'AF_Assemblies') ---")
    if orphan_weapons:
        for weapon_id in sorted(list(orphan_weapons)):
            print(weapon_id)
    else:
        print("No orphan weapons found.")
    
    print("\n\n--- Summary ---")
    print(f"Total defined parts: {len(defined_part_ids)}")
    print(f"Total defined weapons: {len(defined_weapon_ids)}")
    print(f"Total unique equipment IDs used in assemblies: {len(used_equipment_ids)}")
    print(f"Orphan parts found: {len(orphan_parts)}")
    print(f"Orphan weapons found: {len(orphan_weapons)}")
    print("[DEBUG] End of main function reached.")

if __name__ == "__main__":
    main()
    print("\n---------------------------------------------------------------------")
    print("Script execution finished.")
    print("If you encountered an error like 'ModuleNotFoundError: No module named 'openpyxl'',")
    print("you need to install the library by running: pip install openpyxl")
    print("Make sure your Excel file is closed when running the script to avoid read issues.")
    print("---------------------------------------------------------------------") 